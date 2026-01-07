import streamlit as st
import pandas as pd
from openpyxl import load_workbook

st.set_page_config(page_title="Traffic Analytics Dashboard", layout="wide")
st.title("📊 Traffic Performance Dashboard")
st.caption("Point-in-time comparison view across traffic segments")

FILE = "input.xlsx"
SHEET = "Traffic-Status"

wb = load_workbook(FILE)
ws = wb[SHEET]
raw = pd.read_excel(FILE, sheet_name=SHEET, header=None)

# --- find table headers (same logic as agent) ---
headers = [
    i + 1 for i in range(len(raw))
    if "Month" in raw.iloc[i].astype(str).tolist()
    and "Year-2025" in " ".join(raw.iloc[i].astype(str))
]

sections = []

for idx, header in enumerate(headers):
    section_name = f"Segment {idx+1}"

    summary_cell = ws[f"H{header}"].value

    pct_row = header + 1
    while ws.cell(pct_row, 2).value not in ("Total", "% Change"):
        pct_row += 1

    lm = ws.cell(pct_row + 1, 7).value
    yoy = ws.cell(pct_row + 1, 6).value
    ytd = ws.cell(pct_row + 1, 6).value

    sections.append({
        "name": section_name,
        "summary": summary_cell,
        "lm": lm,
        "yoy": yoy,
        "ytd": ytd
    })

# --- UI ---
segment_names = [s["name"] for s in sections]
selected = st.selectbox("Select Traffic Segment", segment_names)

seg = next(s for s in sections if s["name"] == selected)

c1, c2, c3 = st.columns(3)
c1.metric("Latest Month Change", f"{seg['lm']:.2%}" if seg["lm"] else "N/A")
c2.metric("Year-over-Year Change", f"{seg['yoy']:.2%}" if seg["yoy"] else "N/A")
c3.metric("Year-to-Date Change", f"{seg['ytd']:.2%}" if seg["ytd"] else "N/A")

st.subheader("Analytical Summary")
st.write(seg["summary"])

st.caption(
    "All values reflect point-in-time numeric comparisons only "
    "and do not represent sustained trends."
)
