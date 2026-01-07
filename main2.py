
import pandas as pd
from typing import TypedDict, Optional, List
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from dotenv import load_dotenv
from langchain_openai import ChatOpenAI
from langgraph.graph import StateGraph, END
 
# ================= CONFIG =================
load_dotenv()

llm = ChatOpenAI(model="gpt-4o-mini", temperature=0.2)

INPUT_FILE = "input.xlsx"
SHEET_NAME = "Traffic-Status"
SUMMARY_COL = "H"
SUMMARY_ROWS = 13

# ================= HELPERS =================
SECTION_ORDER = [
    "overall website traffic",
    "social media traffic",
    "blog traffic",
    "mobile traffic",
    "desktop traffic",
    "referral traffic",
    "US search traffic",
]



def to_number(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.replace("%", "").replace(",", "").strip()
        if val == "":
            return None
    try:
        return float(val)
    except:
        return None

def set_merged_cell_height(ws, start_row, end_row, text, col_width=110):
    """
    Adjust row height for merged cells based on text length.
    Excel does not auto-resize merged cells.
    """
    chars_per_line = col_width // 7
    total_lines = 0

    for paragraph in text.split("\n"):
        total_lines += max(1, len(paragraph) // chars_per_line + 1)

    height = max(15, total_lines * 15)

    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = height / (end_row - start_row + 1)

def get_section_tone(section_name: str) -> str:
    name = section_name.lower()

    if "total" in name or "overall" in name:
        return (
            "Use an executive-level analytical tone. "
            "Focus on overall performance direction and magnitude. "
            "Avoid operational or granular detail."
        )

    if "blog" in name or "content" in name:
        return (
            "Use a detailed analytical tone. "
            "Emphasize variability, magnitude of change, and consistency over time."
        )

    if "organic" in name or "search" in name:
        return (
            "Use a diagnostic trend-analysis tone. "
            "Emphasize sustained movement, severity, and alignment across timeframes."
        )

    if "engaged" in name:
        return (
            "Use an evaluative quality-focused tone. "
            "Focus on stability, moderation, and directional consistency."
        )

    return (
        "Use a neutral professional analytics tone. "
        "Focus on numeric comparison and observed trends."
    )

def get_segment_label(section_name: str) -> str:
    name = section_name.lower().strip()

    if "total" in name or "overall" in name:
        return "overall website traffic"

    if "social" in name:
        return "social media traffic"

    if "google" in name:
        return "google search traffic"

    if "bing" in name:
        return "bing search traffic"

    if "yahoo" in name:
        return "yahoo search traffic"

    if "blog" in name or "content" in name:
        return "blog traffic"

    if "referral" in name:
        return "referral traffic"

    if "mobile" in name or "smartphone" in name or "tablet" in name:
        return "mobile traffic"

    if "desktop" in name:
        return "desktop traffic"

    if "Us" in name:
        return "US traffic"

    #    FALLBACK
    return "this traffic segment"



# ================= STATE =================
class GAState(TypedDict):
    section_name: str
    months: List[str]
    v24: List[Optional[float]]
    v25: List[Optional[float]]
    yoy: List[Optional[float]]
    lm: List[Optional[float]]
    latest_month: Optional[str]
    insight: str
    plan: str
    planner_debug: str
    summary: str
    ytd_change: Optional[float]
    lm_positive_count: int
    lm_negative_count: int
    avg_abs_lm: Optional[float]
    # global_lm_context: Optional[str]

# ================= AGENT 1: METRICS =================
def metrics_agent(state: GAState) -> GAState:
    yoy, lm = [], []
    latest_month = None

    for i in range(len(state["months"])):
        y = l = None

        if state["v25"][i] is not None and state["v24"][i] not in (None, 0):
            y = (state["v25"][i] - state["v24"][i]) / state["v24"][i]

        if i > 0 and state["v25"][i] is not None and state["v25"][i - 1] not in (None, 0):
            l = (state["v25"][i] - state["v25"][i - 1]) / state["v25"][i - 1]

        yoy.append(y)
        lm.append(l)

        if state["v25"][i] is not None:
            latest_month = state["months"][i]

    

    # 1️ YTD change (sum of current year vs previous year)
    valid_indices = [i for i, v in enumerate(state["v25"]) if v is not None]

    if valid_indices:
        last_idx = max(valid_indices)

        sum_24 = sum(v for v in state["v24"][: last_idx + 1] if v is not None)
        sum_25 = sum(v for v in state["v25"][: last_idx + 1] if v is not None)

        if sum_24 > 0:
            state["ytd_change"] = (sum_25 - sum_24) / sum_24
        else:
            state["ytd_change"] = None
    else:
        state["ytd_change"] = None

    # 2️ LM positive / negative count
    state["lm_positive_count"] = sum(1 for v in lm if v is not None and v > 0)
    state["lm_negative_count"] = sum(1 for v in lm if v is not None and v < 0)

    # 3️ Average absolute LM
    lm_values = [abs(v) for v in lm if v is not None]
    state["avg_abs_lm"] = sum(lm_values) / len(lm_values) if lm_values else None

    # ================= FINAL ASSIGN =================
    state["yoy"] = yoy
    state["lm"] = lm
    state["latest_month"] = latest_month
    

    return state

# ================= AGENT 2: REASONING =================
def reasoning_agent(state: GAState) -> GAState:
    if state["latest_month"] is None:
        state["insight"] = "Insufficient data is available to evaluate recent performance trends."
        return state

    idx = state["months"].index(state["latest_month"])
    lm = state["lm"][idx]
    yoy = state["yoy"][idx]

    prompt = f"""
You are a digital analytics analyst interpreting a comparison between two measured data points.

Rules:
-  Whenever mentioning year-over-year values, you MUST explicitly state that the comparison is for the same month as {state['latest_month']}
- Do NOT use generic phrases like "year-over-year decline" without specifying the month
- Do NOT describe trends, patterns, consistency, acceleration, or momentum
- Do NOT imply long-term or sustained behavior
- Observational and factual tone only
- 5-6 sentences allowed
-You MAY vary phrasing while preserving numeric meaning

Measured Comparisons:
- Adjacent-month change (latest month vs previous month): {round(lm*100,2) if lm is not None else "Not available"}%
- Same-month year-over-year change for {state['latest_month']} compared to the same month last year: {round(yoy*100,2) if yoy is not None else "Not available"}%


Task:
Compare the direction and relative magnitude of these two values using neutral analytical language
"""




    state["insight"] = llm.invoke(prompt).content.strip()
    return state
def planning_agent(state: GAState) -> GAState:
    if state["latest_month"] is None:
        state["plan"] = "Maintain neutral emphasis."
        return state

    idx = state["months"].index(state["latest_month"])
    lm = state["lm"][idx]
    yoy = state["yoy"][idx]

    # 1️ Emphasis planning
    if lm is None or yoy is None:
        state["plan"] = "Maintain neutral emphasis."
    elif abs(lm) > abs(yoy):
        state["plan"] = (
            "Place primary emphasis on the adjacent-month movement. "
            "Mention year-over-year change briefly for context."
        )
    elif abs(yoy) > abs(lm):
        state["plan"] = (
            "Place primary emphasis on the year-over-year comparison. "
            "Mention adjacent-month movement as secondary."
        )
    else:
        state["plan"] = (
            "Maintain balanced emphasis between adjacent-month and year-over-year changes."
        )

    # 2️ Confidence planning
    delta_strength = max(abs(lm or 0), abs(yoy or 0))

    if delta_strength < 0.05:
        confidence = "Use cautious and measured language."
    elif delta_strength < 0.15:
        confidence = "Use moderate analytical language."
    elif delta_strength < 0.30:
        confidence = "Use firm analytical language."
    else:
        confidence = "Use very firm and direct analytical language."


    state["plan"] += " " + confidence

    # 3️ Directional mismatch detection
    if lm is not None and yoy is not None and lm * yoy < 0:
        state["plan"] += (
            " Highlight the directional mismatch between short-term and year-over-year movement."
        )
    debug_lines = []

    debug_lines.append(f"Section: {state['section_name']}")
    debug_lines.append(f"LM: {lm}")
    debug_lines.append(f"YOY: {yoy}")
    debug_lines.append(f"Plan: {state['plan']}")

    state["planner_debug"] = " | ".join(debug_lines)

    return state



# ================= AGENT 3: NARRATIVE =================
def narrative_agent(state: GAState) -> GAState:
    if state["latest_month"] is None:
        state["summary"] = "Insufficient data is available to generate a performance summary."
        return state
    
    plan = state.get("plan", "neutral")
    idx = state["months"].index(state["latest_month"])
    lm = state["lm"][idx]
    yoy = state["yoy"][idx]

    tone_instruction = get_section_tone(state["section_name"])
    segment_label = get_segment_label(state["section_name"])
    prompt = f"""
You are preparing a professional Google Analytics performance commentary
for a specific traffic segment, based strictly on point-in-time numeric comparisons.
The subject of this commentary is: "{segment_label}".
You MUST refer to the data ONLY using this phrase. Do not rename or generalize it.
Tone Guidance:
{tone_instruction}
Strict Rules:
- Do NOT describedelta the data as overall or total traffic unless the section explicitly represents total traffic
- Use ONLY the numeric values provided
- Do NOT use the words: trend, consistency, acceleration, momentum, sustained, long-term
- Do NOT introduce causes, channels, pages, users, assumptions, or influencing factors
- Describe observations based on direct comparison only
- You MUST explicitly mention the distribution of adjacent-month changes (positive vs negative) AND the average absolute adjacent-month magnitude if the values are provided
- Formal report-style language
- You may describe relative magnitude or direction ONLY using the numeric values provided
- Do NOT infer patterns beyond the stated comparisons
Follow the Writing emphasis exactly when ordering and weighting sentences.

Context:
Section: {state['section_name']}
Month: {state['latest_month']}
Writing emphasis: {state['plan']}


Confirmed Metrics (all values are point-in-time and factual):
- Adjacent-month change for {state['latest_month']} compared to the previous month: {round(lm*100,2) if lm is not None else "Not available"}%
- Same-month year-over-year change for {state['latest_month']} compared to the same month last year: {round(yoy*100,2) if yoy is not None else "Not available"}%
- Year-to-date change up to {state['latest_month']} compared to the same period last year: {round(state["ytd_change"]*100,2) if state["ytd_change"] is not None else "Not available"}%
- Number of positive adjacent-month changes across the measured months: {state["lm_positive_count"]}
- Number of negative adjacent-month changes across the measured months: {state["lm_negative_count"]}
- Average absolute adjacent-month change magnitude: {round(state["avg_abs_lm"]*100,2) if state["avg_abs_lm"] is not None else "Not available"}%



Required Structure:
1.Opening sentence stating the latest month’s observed change and magnitude in {segment_label}
2. Sentence comparing the adjacent-month change with the same-month year comparison
3. One sentence explicitly describing the year-to-date change using the numeric value provided
4. One sentence describing the distribution of adjacent-month changes
   (number of positive vs negative months and the average absolute magnitude)
5. Two bullet points restating the key numeric comparisons
6. Closing sentence emphasizing monitoring of future measured changes only
7. Closing sentence emphasizing monitoring of future measured changes only
Mandatory Ending Sentence:
"This interpretation reflects a comparison between measured data points rather than a sustained trend."

"""

    state["summary"] = llm.invoke(prompt).content.strip()
    return state

# ================= BUILD GRAPH =================
graph = StateGraph(GAState)
graph.add_node("metrics", metrics_agent)
graph.add_node("planning", planning_agent)
graph.add_node("reasoning", reasoning_agent)
graph.add_node("narrative", narrative_agent)

graph.set_entry_point("metrics")
graph.add_edge("metrics", "reasoning")
graph.add_edge("reasoning", "planning")
graph.add_edge("planning", "narrative")
graph.add_edge("narrative", END)

agent = graph.compile()

# ================= MAIN =================
def run_agent():
    wb = load_workbook(INPUT_FILE)
    ws = wb[SHEET_NAME]
    raw = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, header=None)

    headers = []

    for i in range(len(raw)):
        row_text = " ".join(raw.iloc[i].astype(str))
        if "Month" in row_text and "Year-2025" in row_text:
            headers.append(i + 1)


    for table_idx, header in enumerate(headers):
        section_name = (
        SECTION_ORDER[table_idx]
        if table_idx < len(SECTION_ORDER)
        else "website traffic")

        months, v23, v24, v25 = [], [], [], []
        r = header + 1

        while r <= ws.max_row:
            m = ws.cell(r, 2).value
            if m is None or str(m).strip() in ("Total", "% Change"):
                break

            months.append(m)
            v23.append(to_number(ws.cell(r, 3).value))
            v24.append(to_number(ws.cell(r, 4).value))
            v25.append(to_number(ws.cell(r, 5).value))
            r += 1

        if not months:
            continue

        state: GAState = {
        "section_name": section_name,
        "months": months,
        "v24": v24,
        "v25": v25,
        "yoy": [],
        "lm": [],
        "latest_month": None,
        "insight": "",
        "plan": "",
        "planner_debug": "",
        "summary": "",
        "ytd_change": None,
        "lm_positive_count": 0,
        "lm_negative_count": 0,
        "avg_abs_lm": None,
        #  "global_lm_context": None
    }


        result = agent.invoke(state)
        print(f"[PLANNER DEBUG] Table {table_idx + 1}: {result['planner_debug']}")
        


        for i, row in enumerate(range(header + 1, header + 1 + len(months))):
            if result["yoy"][i] is not None:
                ws.cell(row, 6).value = result["yoy"][i]
                ws.cell(row, 6).number_format = "0.00%"
            # if result["lm"][i] is not None:
            #     ws.cell(row, 7).value = result["lm"][i]
            #     ws.cell(row, 7).number_format = "0.00%"

        # total_row = header + 1 + len(months)
        # pct_row = total_row + 1
        total_row = r
        pct_row = r + 1
        latest_idx = max( i for i,v in enumerate(v25) if v is not None )
        latest_lm = result["lm"][latest_idx]
        # Write latest LM value ONLY (point-in-time)
        if latest_lm is not None:
            ws.cell(pct_row, 7).value = latest_lm
            ws.cell(pct_row, 7).number_format = "0.00%"

        latest_idx = max( i for i,v in enumerate(v25) if v is not None )
        latest_lm = result["lm"][latest_idx]


        
        # FULL YEAR totals
        sum_23 = sum(v for v in v23 if v is not None)
        sum_24 = sum(v for v in v24 if v is not None)

        # 2025 YTD only
        sum_25 = sum(v25[i] for i in range(latest_idx + 1) if v25[i] is not None)

        # 2024 YTD (same months as 2025)
        sum_24_ytd = sum(v24[i] for i in range(latest_idx + 1) if v24[i] is not None)



        
        

        ws.cell(total_row, 3).value = sum_23
        ws.cell(total_row, 4).value = sum_24
        ws.cell(total_row, 5).value = sum_25

        if sum_23 > 0:
            ws.cell(pct_row, 4).value = (sum_24 - sum_23) / sum_23
            ws.cell(pct_row, 4).number_format = "0.00%"

        if sum_24_ytd > 0:
            ws.cell(pct_row, 6).value = (sum_25 - sum_24_ytd) / sum_24_ytd
            ws.cell(pct_row, 6).number_format = "0.00%"


        start = header
        end = header + SUMMARY_ROWS - 1

        ws.merge_cells(f"H{start}:L{end}")

        cell = ws[f"H{start}"]
        cell.value = result["summary"]
        cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
            horizontal="left"
        )

        set_merged_cell_height(
            ws=ws,
            start_row=start,
            end_row=end,
            text=result["summary"]
        )

    wb.save(INPUT_FILE)
    print(" Totals, % Change, YOY, LM and Narrative successfully generated")

if __name__ == "__main__":
    run_agent()
